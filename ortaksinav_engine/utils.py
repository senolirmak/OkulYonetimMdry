# -*- coding: utf-8 -*-
import re
import warnings

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def normalize_columns(df):
    """DataFrame kolon isimlerini ascii, kucuk harf, alt cizgili formata cevirir."""
    def clean(col):
        col = str(col).strip().lower()
        tr = str.maketrans("şçöğüı ", "scogui_")
        col = col.translate(tr)
        col = re.sub(r"[^a-z0-9_]", "", col)
        return col
    df.columns = [clean(c) for c in df.columns]
    return df


def normalize_sube_cell(s):
    """'9/A, 9/B' gibi alanlari listeye cevirir."""
    if pd.isna(s):
        return []
    s = str(s).upper().replace(" ", "")
    return [x for x in s.split(",") if x]


def safe_sheet_name(name):
    """Excel sheet name icin guvenli isim uretir (max 31 karakter)."""
    name = name.replace("/", "_").replace("\\", "_").replace("?", "_").replace("*", "_")
    name = name.replace("[", "(").replace("]", ")")
    return name[:31]


_EN_GUNLER = {
    0: "Monday", 1: "Tuesday", 2: "Wednesday",
    3: "Thursday", 4: "Friday", 5: "Saturday", 6: "Sunday",
}


def salon_gozetmen_bul(tarih, saat_veya_ders_saati, ss_map: dict) -> dict:
    """
    Verilen tarih/saatte her salon için DersProgrami'nden gözetmen öğretmenini bulur.

    saat_veya_ders_saati: str ("HH:MM") veya DersSaatleri instance
    ss_map  : {"Salon-9_A": SinifSube instance, ...}
    Döndürür: {"Salon-9_A": "ÖĞRETMEN ADI", ...}  — bulunamayanlar boş string.
    """
    from datetime import datetime as _dt
    from dersprogrami.models import DersProgrami
    from okul.models import DersSaatleri as _DS
    from okul.utils import get_aktif_dp_tarihi

    if not ss_map or not tarih or not saat_veya_ders_saati:
        return {k: "" for k in ss_map}

    if isinstance(saat_veya_ders_saati, _DS):
        dp_filter = {"ders_saati": saat_veya_ders_saati}
    else:
        try:
            saat_time = _dt.strptime(saat_veya_ders_saati, "%H:%M").time()
        except (ValueError, TypeError):
            return {k: "" for k in ss_map}
        dp_filter = {"ders_saati__derssaati_baslangic": saat_time}

    aktif_tarih = get_aktif_dp_tarihi()
    if aktif_tarih:
        dp_filter["uygulama_tarihi"] = aktif_tarih

    gun_adi = _EN_GUNLER.get(tarih.weekday(), "")
    dp_map = {
        dp.sinif_sube_id: dp
        for dp in DersProgrami.objects.filter(
            sinif_sube__in=list(ss_map.values()),
            gun=gun_adi,
            **dp_filter,
        ).select_related("ogretmen")
    }
    return {
        salon_adi: (
            dp_map[ss.pk].ogretmen
            if ss.pk in dp_map and dp_map[ss.pk].ogretmen else None
        )
        for salon_adi, ss in ss_map.items()
    }


def iter_rows(file_path):
    """Hem .xls hem .xlsx dosyalarini satirlar halinde dondurur."""
    ext = str(file_path).lower()
    if ext.endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        for i in range(ws.nrows):
            yield ws.row_values(i)
    else:
        wb = load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            yield row
