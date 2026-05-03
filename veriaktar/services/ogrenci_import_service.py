# -*- coding: utf-8 -*-
"""
OgrenciIsleyici – e-Okul OOG01001R020 formatındaki XLS/XLSX dosyasından
öğrenci listesini okuyarak ogrenci.Ogrenci tablosuna yazar.

Format:
- Her sınıf grubunun başında (col0) çok satırlı başlık hücresi:
  "9. Sınıf / A Şubesi" veya "10. Sınıf / B Şubesi" vb.
- Öğrenci satırları: col0 = S.No (sayı > 0), col1 = okulno,
  col4 = adi, col8 = soyadi, col12 = cinsiyet
- Desteklenen formatlar: .XLS / .xls (xlrd), .XLSX / .xlsx (openpyxl)
"""

import re
from datetime import date

import xlrd
from openpyxl import load_workbook


class OgrenciIsleyici:
    SINIF_SUBE_RE = re.compile(
        r"(\d+)\.\s*S[iı]n[iı]f\s*/\s*([A-ZÇĞİÖŞÜa-züçğışö]+)\s*Şubesi",
        re.IGNORECASE,
    )

    def __init__(self, file_path, kullanici=None, dosya_tarihi=None):
        self.file_path = str(file_path)
        self._file_name = str(file_path).split("/")[-1].split("\\")[-1]
        self.kullanici = kullanici
        self.dosya_tarihi = dosya_tarihi
        self._kayitlar = []

    # ------------------------------------------------------------------
    def parse(self) -> list:
        """XLS veya XLSX dosyasını okuyup {sinif, sube, okulno, adi, soyadi, cinsiyet}
        sözlüklerinden oluşan listeyi döner. Uzantı büyük/küçük harf duyarsız."""
        if self.file_path.lower().endswith(".xlsx"):
            return self._parse_xlsx()
        return self._parse_xls()

    def _parse_xls(self) -> list:
        wb = xlrd.open_workbook(self.file_path)
        ws = wb.sheet_by_index(0)

        aktif_sinif = None
        aktif_sube = None
        sonuc = []

        for row_idx in range(ws.nrows):
            val0 = ws.cell_value(row_idx, 0)

            if isinstance(val0, str):
                eslesme = self.SINIF_SUBE_RE.search(val0)
                if eslesme:
                    aktif_sinif = int(eslesme.group(1))
                    aktif_sube = eslesme.group(2).upper().strip()
                continue

            if isinstance(val0, float) and val0 > 0 and aktif_sinif is not None:
                try:
                    okulno = str(ws.cell_value(row_idx, 1)).strip()
                    if okulno.endswith(".0"):
                        okulno = okulno[:-2]
                    adi = str(ws.cell_value(row_idx, 4)).strip().upper()
                    soyadi = str(ws.cell_value(row_idx, 8)).strip().upper()
                    cinsiyet_raw = str(ws.cell_value(row_idx, 12)).strip().lower()
                    cinsiyet = "E" if "erkek" in cinsiyet_raw else "K"

                    if not okulno or not adi or not soyadi:
                        continue

                    sonuc.append({
                        "sinif": aktif_sinif,
                        "sube": aktif_sube,
                        "okulno": okulno,
                        "adi": adi,
                        "soyadi": soyadi,
                        "cinsiyet": cinsiyet,
                    })
                except (IndexError, ValueError):
                    continue

        self._kayitlar = sonuc
        return sonuc

    def _parse_xlsx(self) -> list:
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.active

        aktif_sinif = None
        aktif_sube = None
        sonuc = []

        for row in ws.iter_rows(values_only=True):
            val0 = row[0]

            # Başlık hücresi: string içinde "X. Sınıf / Y Şubesi"
            if val0 is not None and not isinstance(val0, (int, float)):
                eslesme = self.SINIF_SUBE_RE.search(str(val0))
                if eslesme:
                    aktif_sinif = int(eslesme.group(1))
                    aktif_sube = eslesme.group(2).upper().strip()
                continue

            # Öğrenci satırı: col0 sayısal S.No > 0
            if isinstance(val0, (int, float)) and val0 > 0 and aktif_sinif is not None:
                try:
                    okulno = str(row[1]).strip() if row[1] is not None else ""
                    if okulno.endswith(".0"):
                        okulno = okulno[:-2]
                    adi = str(row[4]).strip().upper() if row[4] is not None else ""
                    soyadi = str(row[8]).strip().upper() if row[8] is not None else ""
                    cinsiyet_raw = str(row[12]).strip().lower() if row[12] is not None else ""
                    cinsiyet = "E" if "erkek" in cinsiyet_raw else "K"

                    if not okulno or not adi or not soyadi:
                        continue

                    sonuc.append({
                        "sinif": aktif_sinif,
                        "sube": aktif_sube,
                        "okulno": okulno,
                        "adi": adi,
                        "soyadi": soyadi,
                        "cinsiyet": cinsiyet,
                    })
                except (IndexError, ValueError):
                    continue

        wb.close()
        self._kayitlar = sonuc
        return sonuc

    # ------------------------------------------------------------------
    def veritabanina_yaz(self) -> dict:
        """parse() sonucunu ogrenci.Ogrenci tablosuna yazar.
        tckimlikno ve dogumtarihi eksik olduğu için yer tutucu değer kullanılır."""
        from ogrenci.models import Ogrenci

        yeni = 0
        guncellenen = 0
        hatali = 0

        for kayit in self._kayitlar:
            okulno = kayit["okulno"]
            tckimlikno_placeholder = okulno.zfill(11)
            try:
                obj, created = Ogrenci.objects.update_or_create(
                    okulno=okulno,
                    defaults={
                        "sinif": kayit["sinif"],
                        "sube": kayit["sube"],
                        "adi": kayit["adi"],
                        "soyadi": kayit["soyadi"],
                        "cinsiyet": kayit["cinsiyet"],
                        "tckimlikno": tckimlikno_placeholder,
                        "dogumtarihi": date(2000, 1, 1),
                    },
                )
                if created:
                    yeni += 1
                else:
                    guncellenen += 1
            except Exception:
                hatali += 1

        return {"yeni": yeni, "guncellenen": guncellenen, "hatali": hatali}

    # ------------------------------------------------------------------
    def _aktar_gecmisi_kaydet(self, status):
        from okul.models import VeriAktarimGecmisi

        durum = "basarili" if not status.get("hatali") else "kismi"

        VeriAktarimGecmisi.objects.create(
            dosya_turu="ogrenci_listesi",
            dosya_adi=self._file_name,
            dosya_tarihi=self.dosya_tarihi,
            kullanici=self.kullanici,
            kayit_sayisi=status.get("yeni", 0) + status.get("guncellenen", 0),
            hata_sayisi=status.get("hatali", 0),
            durum=durum,
        )

    def calistir(self) -> dict:
        self.parse()
        status = self.veritabanina_yaz()
        self._aktar_gecmisi_kaydet(status)
        return status
